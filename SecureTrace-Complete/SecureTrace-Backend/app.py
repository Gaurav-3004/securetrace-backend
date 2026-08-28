import io
import os
import hashlib
import secrets
import smtplib
import random
import string
from datetime import datetime, timedelta
from email.mime.text import MIMEText
from functools import wraps

import jwt
import psycopg2
import psycopg2.extras
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from flask import Flask, request, jsonify, g, send_file

app = Flask(__name__)

# ---------------- Config (set these as environment variables on your host) ----------------
JWT_SECRET = os.environ.get("JWT_SECRET", "change-this-secret-in-production")
JWT_EXPIRY_DAYS = 7

GMAIL_ADDRESS = os.environ.get("GMAIL_ADDRESS", "")
GMAIL_APP_PASSWORD = os.environ.get("GMAIL_APP_PASSWORD", "")

FAILED_ATTEMPT_THRESHOLD = 5
OTP_VALID_MINUTES = 5

# Postgres connection string, e.g. postgresql://user:pass@host/dbname
# Get this from Neon.tech (or Render Postgres, Supabase, etc.) and set it as
# the DATABASE_URL environment variable on your host.
DATABASE_URL = os.environ.get("DATABASE_URL", "")

DATE_FORMAT = "%d %b %Y, %I:%M:%S %p"


# ---------------- Database ----------------

def get_db():
    if "db" not in g:
        if not DATABASE_URL:
            raise RuntimeError(
                "DATABASE_URL is not set. Add your Postgres connection string as an "
                "environment variable on your hosting provider."
            )
        g.db = psycopg2.connect(DATABASE_URL, cursor_factory=psycopg2.extras.RealDictCursor)
    return g.db


@app.teardown_appcontext
def close_db(exception=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db():
    if not DATABASE_URL:
        print("[WARN] DATABASE_URL not set — skipping DB init. Set it and redeploy.")
        return

    conn = psycopg2.connect(DATABASE_URL)
    c = conn.cursor()

    c.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id SERIAL PRIMARY KEY,
            username TEXT UNIQUE NOT NULL,
            email TEXT NOT NULL,
            password_hash TEXT NOT NULL,
            salt TEXT NOT NULL,
            is_admin INTEGER DEFAULT 0,
            is_active INTEGER DEFAULT 1,
            failed_attempts INTEGER DEFAULT 0,
            created_at TEXT
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS login_activity (
            id SERIAL PRIMARY KEY,
            user_id INTEGER,
            username TEXT,
            device_name TEXT,
            device_model TEXT,
            device_type TEXT,
            operating_system TEXT,
            app_version TEXT,
            login_status TEXT,
            login_time TEXT,
            logout_time TEXT,
            is_active INTEGER DEFAULT 0,
            approx_location TEXT
        )
    """)

    # Migration: add the column if this table already existed from before this feature
    c.execute("""
        DO $$
        BEGIN
            IF NOT EXISTS (
                SELECT 1 FROM information_schema.columns
                WHERE table_name='login_activity' AND column_name='approx_location'
            ) THEN
                ALTER TABLE login_activity ADD COLUMN approx_location TEXT;
            END IF;
        END $$;
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS security_alerts (
            id SERIAL PRIMARY KEY,
            user_id INTEGER,
            alert_type TEXT,
            description TEXT,
            created_at TEXT,
            is_read INTEGER DEFAULT 0
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS otp_codes (
            id SERIAL PRIMARY KEY,
            email TEXT,
            code TEXT,
            purpose TEXT,
            payload TEXT,
            expires_at TEXT
        )
    """)

    # Seed default admin: admin / admin123
    c.execute("SELECT id FROM users WHERE username = 'admin'")
    if c.fetchone() is None:
        salt = generate_salt()
        c.execute(
            "INSERT INTO users (username, email, password_hash, salt, is_admin, is_active, created_at) "
            "VALUES (%s, %s, %s, %s, 1, 1, %s)",
            ("admin", "baviskarg604@gmail.com", hash_password("admin123", salt), salt, now_str())
        )
    else:
        # One-time fix: if the admin account was already created with the old
        # placeholder email (which can never receive OTPs), correct it now.
        c.execute(
            "UPDATE users SET email = %s WHERE username = 'admin' AND email = %s",
            ("baviskarg604@gmail.com", "admin@securetrace.app")
        )

    conn.commit()
    conn.close()


# ---------------- Helpers ----------------

def now_str():
    return datetime.utcnow().strftime(DATE_FORMAT)


def generate_salt():
    return secrets.token_hex(16)


def hash_password(password, salt):
    return hashlib.sha256((salt + password).encode()).hexdigest()


def generate_otp():
    return "".join(random.choices(string.digits, k=6))


def send_email(to_email, subject, body):
    if not GMAIL_ADDRESS or not GMAIL_APP_PASSWORD:
        print(f"[EMAIL NOT CONFIGURED] Would send to {to_email}: {body}")
        return False
    try:
        msg = MIMEText(body)
        msg["Subject"] = subject
        msg["From"] = GMAIL_ADDRESS
        msg["To"] = to_email

        with smtplib.SMTP("smtp.gmail.com", 587) as server:
            server.starttls()
            server.login(GMAIL_ADDRESS, GMAIL_APP_PASSWORD)
            server.sendmail(GMAIL_ADDRESS, [to_email], msg.as_string())
        return True
    except Exception as e:
        print(f"[EMAIL ERROR] {e}")
        return False


def create_token(user_id, username, is_admin):
    payload = {
        "user_id": user_id,
        "username": username,
        "is_admin": is_admin,
        "exp": datetime.utcnow() + timedelta(days=JWT_EXPIRY_DAYS)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm="HS256")


def require_auth(admin_only=False):
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            auth_header = request.headers.get("Authorization", "")
            if not auth_header.startswith("Bearer "):
                return jsonify({"error": "Missing or invalid Authorization header"}), 401
            token = auth_header.replace("Bearer ", "")
            try:
                payload = jwt.decode(token, JWT_SECRET, algorithms=["HS256"])
            except jwt.ExpiredSignatureError:
                return jsonify({"error": "Session expired, please log in again"}), 401
            except jwt.InvalidTokenError:
                return jsonify({"error": "Invalid token"}), 401

            if admin_only and not payload.get("is_admin"):
                return jsonify({"error": "Admin access required"}), 403

            g.current_user = payload
            return f(*args, **kwargs)
        return wrapper
    return decorator


def mask_email(email):
    try:
        name, domain = email.split("@")
        visible = name[:2]
        return f"{visible}{'*' * max(len(name) - 2, 1)}@{domain}"
    except Exception:
        return email


def create_alert(db, user_id, alert_type, description):
    db.execute(
        "INSERT INTO security_alerts (user_id, alert_type, description, created_at, is_read) "
        "VALUES (%s, %s, %s, %s, 0)",
        (user_id, alert_type, description, now_str())
    )


def get_client_ip():
    """Best-effort client IP, accounting for Render's reverse proxy."""
    forwarded = request.headers.get("X-Forwarded-For", "")
    if forwarded:
        return forwarded.split(",")[0].strip()
    return request.remote_addr or ""


def get_approx_location(ip):
    """
    Resolves an IP address to a city-level location only (e.g. "Mumbai, IN").
    Never resolves GPS coordinates or a precise address. Fails silently
    (returns None) so a slow/unavailable geolocation service never blocks login.
    """
    if not ip or ip in ("127.0.0.1", "localhost", "::1"):
        return None
    try:
        resp = requests.get(
            f"http://ip-api.com/json/{ip}",
            params={"fields": "status,city,country"},
            timeout=3
        )
        data = resp.json()
        if data.get("status") == "success":
            city = data.get("city")
            country = data.get("country")
            if city and country:
                return f"{city}, {country}"
            return city or country
    except Exception:
        pass
    return None



def with_duration(row):
    """Adds a computed duration_minutes field to an activity row dict."""
    row = dict(row)
    login_time = row.get("login_time")
    logout_time = row.get("logout_time")
    duration = None
    try:
        if login_time:
            start = datetime.strptime(login_time, DATE_FORMAT)
            if logout_time:
                end = datetime.strptime(logout_time, DATE_FORMAT)
            elif row.get("is_active"):
                end = datetime.utcnow()
            else:
                end = None
            if end:
                duration = max(int((end - start).total_seconds() // 60), 0)
    except Exception:
        duration = None
    row["duration_minutes"] = duration
    return row


def filter_by_days(rows, days):
    """Keeps only rows whose login_time falls within the last `days` days.
    If days is None/0, returns all rows unchanged."""
    if not days:
        return rows
    cutoff = datetime.utcnow() - timedelta(days=days)
    result = []
    for r in rows:
        try:
            t = datetime.strptime(r.get("login_time", ""), DATE_FORMAT)
            if t >= cutoff:
                result.append(r)
        except Exception:
            result.append(r)  # keep rows we can't parse rather than silently drop them
    return result


def build_activity_excel(rows, sheet_title="Activity"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    include_location = any("approx_location" in r for r in rows)

    headers = [
        "ID", "Username", "Device Name", "Device Model", "Device Type",
        "Operating System", "App Version", "Status", "Login Time",
        "Logout Time", "Duration (minutes)"
    ]
    if include_location:
        headers.append("Approx. Location (from IP)")
    ws.append(headers)

    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for r in rows:
        row_values = [
            r.get("id"),
            r.get("username"),
            r.get("device_name"),
            r.get("device_model"),
            r.get("device_type"),
            r.get("operating_system"),
            r.get("app_version"),
            r.get("login_status"),
            r.get("login_time"),
            r.get("logout_time") or "",
            r.get("duration_minutes") if r.get("duration_minutes") is not None else "",
        ]
        if include_location:
            row_values.append(r.get("approx_location") or "")
        ws.append(row_values)

    widths = [6, 16, 18, 16, 14, 16, 12, 10, 22, 22, 16, 22]
    for i, w in enumerate(widths[:len(headers)], start=1):
        col_letter = chr(64 + i) if i <= 26 else "A"
        ws.column_dimensions[col_letter].width = w

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer




@app.route("/api/register", methods=["POST"])
def register():
    data = request.get_json(force=True)
    username = (data.get("username") or "").strip()
    email = (data.get("email") or "").strip()
    password = (data.get("password") or "").strip()

    if not username or not email or not password:
        return jsonify({"error": "All fields are required"}), 400
    if len(password) < 6:
        return jsonify({"error": "Password must be at least 6 characters"}), 400

    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT id FROM users WHERE username = %s", (username,))
    if cur.fetchone():
        return jsonify({"error": "Username already taken"}), 409

    otp = generate_otp()
    expires_at = (datetime.utcnow() + timedelta(minutes=OTP_VALID_MINUTES)).isoformat()
    payload = f"{username}||{email}||{password}"

    cur.execute("DELETE FROM otp_codes WHERE email = %s AND purpose = 'register'", (email,))
    cur.execute(
        "INSERT INTO otp_codes (email, code, purpose, payload, expires_at) VALUES (%s, %s, 'register', %s, %s)",
        (email, otp, payload, expires_at)
    )
    db.commit()

    sent = send_email(email, "Your SecureTrace verification code",
                       f"Your SecureTrace verification code is: {otp}\n\nThis code expires in {OTP_VALID_MINUTES} minutes.")

    return jsonify({"message": "OTP sent", "email_sent": sent})


# ---------------- Auth: Login (step 1 — password check, sends OTP) ----------------

@app.route("/api/login", methods=["POST"])
def login():
    data = request.get_json(force=True)
    username = (data.get("username") or "").strip()
    password = (data.get("password") or "").strip()
    device_name = data.get("device_name", "Unknown device")
    device_model = data.get("device_model", "")
    device_type = data.get("device_type", "Phone")
    operating_system = data.get("operating_system", "")
    app_version = data.get("app_version", "1.0")

    if not username or not password:
        return jsonify({"error": "Username and password are required"}), 400

    db = get_db()
    cur = db.cursor()
    cur.execute(
        "SELECT id, email, password_hash, salt, is_admin, is_active, failed_attempts FROM users WHERE username = %s",
        (username,)
    )
    user = cur.fetchone()

    if user is None:
        return jsonify({"error": "Invalid username or password"}), 401

    if not user["is_active"]:
        return jsonify({"error": "This account is locked or disabled"}), 403

    if hash_password(password, user["salt"]) != user["password_hash"]:
        new_fail_count = user["failed_attempts"] + 1
        cur.execute(
            "INSERT INTO login_activity (user_id, username, device_name, device_model, device_type, "
            "operating_system, app_version, login_status, login_time, is_active) "
            "VALUES (%s, %s, %s, %s, %s, %s, %s, 'FAILED', %s, 0)",
            (user["id"], username, device_name, device_model, device_type, operating_system, app_version, now_str())
        )

        if new_fail_count >= FAILED_ATTEMPT_THRESHOLD:
            cur.execute("UPDATE users SET failed_attempts = %s, is_active = 0 WHERE id = %s", (new_fail_count, user["id"]))
            create_alert(cur, user["id"], "MULTIPLE_FAILED_ATTEMPTS",
                         f"{new_fail_count} consecutive failed login attempts detected. Account has been locked.")
            db.commit()
            return jsonify({"error": "Too many failed attempts. Account has been locked."}), 403
        else:
            cur.execute("UPDATE users SET failed_attempts = %s WHERE id = %s", (new_fail_count, user["id"]))
            if new_fail_count >= 3:
                create_alert(cur, user["id"], "FAILED_LOGIN", f"{new_fail_count} failed login attempts on your account.")
            db.commit()

        return jsonify({"error": "Invalid username or password"}), 401

    # Password correct — issue OTP, defer session/activity logging to verify step
    otp = generate_otp()
    expires_at = (datetime.utcnow() + timedelta(minutes=OTP_VALID_MINUTES)).isoformat()
    payload = (
        f"{user['id']}||{username}||{user['is_admin']}||{device_name}||{device_model}||"
        f"{device_type}||{operating_system}||{app_version}||{user['failed_attempts']}"
    )

    cur.execute("DELETE FROM otp_codes WHERE email = %s AND purpose = 'login'", (user["email"],))
    cur.execute(
        "INSERT INTO otp_codes (email, code, purpose, payload, expires_at) VALUES (%s, %s, 'login', %s, %s)",
        (user["email"], otp, payload, expires_at)
    )
    db.commit()

    sent = send_email(user["email"], "Your SecureTrace login code",
                       f"Your SecureTrace login code is: {otp}\n\nThis code expires in {OTP_VALID_MINUTES} minutes.")

    return jsonify({
        "message": "OTP sent",
        "email_sent": sent,
        "masked_email": mask_email(user["email"]),
        "email": user["email"]
    })


# ---------------- Auth: Verify OTP (completes register OR login) ----------------

@app.route("/api/verify-otp", methods=["POST"])
def verify_otp():
    data = request.get_json(force=True)
    email = (data.get("email") or "").strip()
    code = (data.get("otp") or "").strip()
    purpose = (data.get("purpose") or "").strip()

    db = get_db()
    cur = db.cursor()
    cur.execute(
        "SELECT * FROM otp_codes WHERE email = %s AND purpose = %s ORDER BY id DESC LIMIT 1",
        (email, purpose)
    )
    row = cur.fetchone()

    if row is None:
        return jsonify({"error": "No OTP request found. Please try again."}), 400

    if datetime.utcnow() > datetime.fromisoformat(row["expires_at"]):
        cur.execute("DELETE FROM otp_codes WHERE id = %s", (row["id"],))
        db.commit()
        return jsonify({"error": "Code expired. Please request a new one."}), 400

    if row["code"] != code:
        return jsonify({"error": "Incorrect code"}), 400

    cur.execute("DELETE FROM otp_codes WHERE id = %s", (row["id"],))

    if purpose == "register":
        username, reg_email, password = row["payload"].split("||")
        cur.execute("SELECT id FROM users WHERE username = %s", (username,))
        if cur.fetchone():
            db.commit()
            return jsonify({"error": "That username was just taken. Please try again."}), 409

        salt = generate_salt()
        cur.execute(
            "INSERT INTO users (username, email, password_hash, salt, is_admin, is_active, created_at) "
            "VALUES (%s, %s, %s, %s, 0, 1, %s)",
            (username, reg_email, hash_password(password, salt), salt, now_str())
        )
        db.commit()
        return jsonify({"message": "Account created successfully"})

    elif purpose == "login":
        parts = row["payload"].split("||")
        user_id, username, is_admin, device_name, device_model, device_type, operating_system, app_version, prev_failed = parts
        user_id = int(user_id)
        is_admin = int(is_admin)
        prev_failed = int(prev_failed)

        cur.execute("UPDATE users SET failed_attempts = 0 WHERE id = %s", (user_id,))

        if prev_failed >= 3:
            create_alert(cur, user_id, "UNUSUAL_PATTERN",
                         f"Successful login after {prev_failed} failed attempts. If this wasn't you, change your password.")

        cur.execute(
            "SELECT id FROM login_activity WHERE user_id = %s AND device_name = %s AND device_model = %s "
            "AND login_status = 'SUCCESS' LIMIT 1",
            (user_id, device_name, device_model)
        )
        is_new_device = cur.fetchone() is None

        approx_location = get_approx_location(get_client_ip())

        cur.execute(
            "INSERT INTO login_activity (user_id, username, device_name, device_model, device_type, "
            "operating_system, app_version, login_status, login_time, is_active, approx_location) "
            "VALUES (%s, %s, %s, %s, %s, %s, %s, 'SUCCESS', %s, 1, %s) RETURNING id",
            (user_id, username, device_name, device_model, device_type, operating_system, app_version, now_str(), approx_location)
        )
        log_id = cur.fetchone()["id"]

        if is_new_device:
            create_alert(cur, user_id, "NEW_DEVICE", f"New sign-in detected from {device_name} ({operating_system}).")

        db.commit()

        token = create_token(user_id, username, bool(is_admin))
        return jsonify({
            "message": "Login successful",
            "token": token,
            "user": {"id": user_id, "username": username, "is_admin": bool(is_admin)},
            "log_id": log_id
        })

    else:
        db.commit()
        return jsonify({"error": "Unknown OTP purpose"}), 400


@app.route("/api/resend-otp", methods=["POST"])
def resend_otp():
    data = request.get_json(force=True)
    email = (data.get("email") or "").strip()
    purpose = (data.get("purpose") or "").strip()

    db = get_db()
    cur = db.cursor()
    cur.execute(
        "SELECT * FROM otp_codes WHERE email = %s AND purpose = %s ORDER BY id DESC LIMIT 1",
        (email, purpose)
    )
    row = cur.fetchone()
    if row is None:
        return jsonify({"error": "No pending request to resend"}), 400

    otp = generate_otp()
    expires_at = (datetime.utcnow() + timedelta(minutes=OTP_VALID_MINUTES)).isoformat()
    cur.execute("UPDATE otp_codes SET code = %s, expires_at = %s WHERE id = %s", (otp, expires_at, row["id"]))
    db.commit()

    sent = send_email(email, "Your SecureTrace verification code",
                       f"Your SecureTrace verification code is: {otp}\n\nThis code expires in {OTP_VALID_MINUTES} minutes.")
    return jsonify({"message": "OTP resent", "email_sent": sent})


# ---------------- Password reset / change ----------------

@app.route("/api/forgot-password", methods=["POST"])
def forgot_password():
    data = request.get_json(force=True)
    username = (data.get("username") or "").strip()
    email = (data.get("email") or "").strip()
    new_password = (data.get("new_password") or "").strip()

    if len(new_password) < 6:
        return jsonify({"error": "Password must be at least 6 characters"}), 400

    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT id FROM users WHERE username = %s AND email = %s", (username, email))
    user = cur.fetchone()
    if user is None:
        return jsonify({"error": "Username and email don't match our records"}), 404

    salt = generate_salt()
    cur.execute(
        "UPDATE users SET password_hash = %s, salt = %s, failed_attempts = 0 WHERE id = %s",
        (hash_password(new_password, salt), salt, user["id"])
    )
    create_alert(cur, user["id"], "PASSWORD_CHANGED", "Your password was reset successfully.")
    db.commit()
    return jsonify({"message": "Password reset successful"})


@app.route("/api/change-password", methods=["POST"])
@require_auth()
def change_password():
    data = request.get_json(force=True)
    old_password = (data.get("old_password") or "").strip()
    new_password = (data.get("new_password") or "").strip()
    user_id = g.current_user["user_id"]

    if len(new_password) < 6:
        return jsonify({"error": "New password must be at least 6 characters"}), 400

    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT password_hash, salt FROM users WHERE id = %s", (user_id,))
    user = cur.fetchone()
    if hash_password(old_password, user["salt"]) != user["password_hash"]:
        return jsonify({"error": "Current password is incorrect"}), 400

    new_salt = generate_salt()
    cur.execute(
        "UPDATE users SET password_hash = %s, salt = %s WHERE id = %s",
        (hash_password(new_password, new_salt), new_salt, user_id)
    )
    create_alert(cur, user_id, "PASSWORD_CHANGED", "Your account password was changed.")
    db.commit()
    return jsonify({"message": "Password updated successfully"})


# ---------------- Profile ----------------

@app.route("/api/me", methods=["GET"])
@require_auth()
def me():
    user_id = g.current_user["user_id"]
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT username, email, created_at FROM users WHERE id = %s", (user_id,))
    user = cur.fetchone()
    if user is None:
        return jsonify({"error": "User not found"}), 404
    return jsonify({"username": user["username"], "email": user["email"], "created_at": user["created_at"]})


# ---------------- User dashboard ----------------

@app.route("/api/dashboard", methods=["GET"])
@require_auth()
def dashboard():
    user_id = g.current_user["user_id"]
    db = get_db()
    cur = db.cursor()

    cur.execute("SELECT COUNT(*) c FROM login_activity WHERE user_id = %s AND login_status = 'SUCCESS'", (user_id,))
    total_logins = cur.fetchone()["c"]

    cur.execute(
        "SELECT login_time FROM login_activity WHERE user_id = %s AND login_status = 'SUCCESS' "
        "ORDER BY id DESC LIMIT 1", (user_id,)
    )
    last_login_row = cur.fetchone()
    last_login = last_login_row["login_time"] if last_login_row else None

    cur.execute(
        "SELECT COUNT(DISTINCT device_name || device_model) c FROM login_activity WHERE user_id = %s AND is_active = 1",
        (user_id,)
    )
    active_devices = cur.fetchone()["c"]

    cur.execute("SELECT COUNT(*) c FROM security_alerts WHERE user_id = %s AND is_read = 0", (user_id,))
    unread_alerts = cur.fetchone()["c"]

    cur.execute("SELECT * FROM login_activity WHERE user_id = %s ORDER BY id DESC LIMIT 5", (user_id,))
    recent = [with_duration(r) for r in cur.fetchall()]

    return jsonify({
        "total_logins": total_logins,
        "last_login": last_login,
        "active_devices": active_devices,
        "unread_alerts": unread_alerts,
        "recent_activity": recent
    })


@app.route("/api/activity", methods=["GET"])
@require_auth()
def activity_history():
    user_id = g.current_user["user_id"]
    days = request.args.get("days", type=int)
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT * FROM login_activity WHERE user_id = %s ORDER BY id DESC LIMIT 500", (user_id,))
    rows = [with_duration(r) for r in cur.fetchall()]
    rows = filter_by_days(rows, days)
    return jsonify({"activity": rows})


@app.route("/api/activity/export", methods=["GET"])
@require_auth()
def export_my_activity():
    user_id = g.current_user["user_id"]
    username = g.current_user["username"]
    days = request.args.get("days", type=int)
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT * FROM login_activity WHERE user_id = %s ORDER BY id DESC LIMIT 5000", (user_id,))
    rows = [with_duration(r) for r in cur.fetchall()]
    rows = filter_by_days(rows, days)

    buffer = build_activity_excel(rows, sheet_title="My Activity")
    filename = f"securetrace_{username}_activity.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/api/alerts", methods=["GET"])
@require_auth()
def alerts():
    user_id = g.current_user["user_id"]
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT * FROM security_alerts WHERE user_id = %s ORDER BY id DESC", (user_id,))
    rows = [dict(r) for r in cur.fetchall()]
    cur.execute("UPDATE security_alerts SET is_read = 1 WHERE user_id = %s", (user_id,))
    db.commit()
    return jsonify({"alerts": rows})


@app.route("/api/logout", methods=["POST"])
@require_auth()
def logout():
    data = request.get_json(force=True)
    log_id = data.get("log_id")
    db = get_db()
    cur = db.cursor()
    cur.execute(
        "UPDATE login_activity SET logout_time = %s, is_active = 0 WHERE id = %s",
        (now_str(), log_id)
    )
    db.commit()
    return jsonify({"message": "Logged out"})


# ---------------- Admin ----------------

@app.route("/api/admin/stats", methods=["GET"])
@require_auth(admin_only=True)
def admin_stats():
    db = get_db()
    cur = db.cursor()

    def count(query, params=()):
        cur.execute(query, params)
        return cur.fetchone()["c"]

    return jsonify({
        "total_users": count("SELECT COUNT(*) c FROM users"),
        "total_activities": count("SELECT COUNT(*) c FROM login_activity"),
        "failed_logins": count("SELECT COUNT(*) c FROM login_activity WHERE login_status = 'FAILED'"),
        "success_logins": count("SELECT COUNT(*) c FROM login_activity WHERE login_status = 'SUCCESS'"),
        "total_alerts": count("SELECT COUNT(*) c FROM security_alerts"),
        "active_users": count("SELECT COUNT(*) c FROM users WHERE is_active = 1"),
        "disabled_users": count("SELECT COUNT(*) c FROM users WHERE is_active = 0"),
    })


@app.route("/api/admin/users", methods=["GET"])
@require_auth(admin_only=True)
def admin_users():
    search = request.args.get("search", "").strip()
    db = get_db()
    cur = db.cursor()
    if search:
        like = f"%{search}%"
        cur.execute(
            "SELECT id, username, email, is_admin, is_active, created_at FROM users "
            "WHERE username ILIKE %s OR email ILIKE %s ORDER BY id DESC",
            (like, like)
        )
    else:
        cur.execute("SELECT id, username, email, is_admin, is_active, created_at FROM users ORDER BY id DESC")
    return jsonify({"users": [dict(r) for r in cur.fetchall()]})


@app.route("/api/admin/users/<int:user_id>/toggle", methods=["POST"])
@require_auth(admin_only=True)
def admin_toggle_user(user_id):
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT is_active FROM users WHERE id = %s", (user_id,))
    user = cur.fetchone()
    if user is None:
        return jsonify({"error": "User not found"}), 404

    new_status = 0 if user["is_active"] else 1
    cur.execute("UPDATE users SET is_active = %s, failed_attempts = 0 WHERE id = %s", (new_status, user_id))
    create_alert(
        cur, user_id, "ACCOUNT_UPDATE",
        "Your account was re-enabled by an administrator." if new_status else
        "Your account was disabled by an administrator."
    )
    db.commit()
    return jsonify({"message": "Updated", "is_active": bool(new_status)})


def strip_location(rows):
    """Removes approx_location before returning data to admin views — a user's
    approximate login location is visible only to that user themselves."""
    for r in rows:
        r.pop("approx_location", None)
    return rows


@app.route("/api/admin/activity", methods=["GET"])
@require_auth(admin_only=True)
def admin_activity():
    days = request.args.get("days", type=int)
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT * FROM login_activity ORDER BY id DESC LIMIT 500")
    rows = [with_duration(r) for r in cur.fetchall()]
    rows = filter_by_days(rows, days)
    rows = strip_location(rows)
    return jsonify({"activity": rows})


@app.route("/api/admin/activity/export", methods=["GET"])
@require_auth(admin_only=True)
def export_admin_activity():
    days = request.args.get("days", type=int)
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT * FROM login_activity ORDER BY id DESC LIMIT 10000")
    rows = [with_duration(r) for r in cur.fetchall()]
    rows = filter_by_days(rows, days)
    rows = strip_location(rows)

    buffer = build_activity_excel(rows, sheet_title="All Users Activity")
    return send_file(
        buffer,
        as_attachment=True,
        download_name="securetrace_all_users_activity.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )



@app.route("/api/admin/analytics", methods=["GET"])
@require_auth(admin_only=True)
def admin_analytics():
    db = get_db()
    cur = db.cursor()
    trend = []
    for i in range(6, -1, -1):
        day = (datetime.utcnow() - timedelta(days=i))
        day_key = day.strftime("%d %b %Y")
        label = day.strftime("%d %b")
        cur.execute(
            "SELECT COUNT(*) c FROM login_activity WHERE login_status = 'SUCCESS' AND login_time LIKE %s",
            (f"{day_key}%",)
        )
        success = cur.fetchone()["c"]
        cur.execute(
            "SELECT COUNT(*) c FROM login_activity WHERE login_status = 'FAILED' AND login_time LIKE %s",
            (f"{day_key}%",)
        )
        failed = cur.fetchone()["c"]
        trend.append({"label": label, "success": success, "failed": failed})

    cur.execute("SELECT device_type, COUNT(*) c FROM login_activity GROUP BY device_type ORDER BY c DESC")
    devices = cur.fetchall()

    return jsonify({
        "trend": trend,
        "device_distribution": [{"type": r["device_type"] or "Unknown", "count": r["c"]} for r in devices]
    })


@app.route("/", methods=["GET"])
def health_check():
    return jsonify({"status": "SecureTrace API is running"})


init_db()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
